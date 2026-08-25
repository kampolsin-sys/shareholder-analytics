import streamlit as st
import pandas as pd
import database as db
import io
import os
import base64

@st.cache_data
def get_base64_of_bin_file(bin_file):
    with open(bin_file, 'rb') as f:
        data = f.read()
    return base64.b64encode(data).decode()

def set_background(png_file):
    if os.path.exists(png_file):
        bin_str = get_base64_of_bin_file(png_file)
        page_bg_img = f'''
        <style>
        .stApp {{
            background-image: linear-gradient(rgba(255,255,255,0.85), rgba(255,255,255,0.85)), url("data:image/png;base64,{bin_str}");
            background-size: cover;
            background-position: center;
            background-repeat: no-repeat;
            background-attachment: fixed;
        }}
        </style>
        '''
        st.markdown(page_bg_img, unsafe_allow_html=True)

BG_IMAGE_PATH = "bg.png"

try:
    if "view" in st.query_params and st.query_params["view"] == "report":
        st.set_page_config(page_title="Shareable Report", layout="wide", initial_sidebar_state="collapsed")
        set_background(BG_IMAGE_PATH)
        st.markdown("""
            <style>
                #MainMenu {visibility: hidden;}
                footer {visibility: hidden;}
                header {visibility: hidden;}
                .block-container {padding-top: 1rem; max-width: 95%;}
                [data-testid="collapsedControl"] {display: none;}
            </style>
        """, unsafe_allow_html=True)
        if os.path.exists("shareable_report.html"):
            with open("shareable_report.html", "r", encoding="utf-8") as f:
                html_content = f.read()
                if hasattr(st, 'html'):
                    st.html(html_content)
                else:
                    st.components.v1.html(html_content, height=800, scrolling=True)
        else:
            st.error("ไม่พบข้อมูลรายงาน กรุณากลับไปหน้าหลักแล้วกด 'ประมวลผลเปรียบเทียบ' อีกครั้ง")
        st.stop()
except Exception:
    pass

import auth

# Initialize DB to ensure users table and default admin exists
db.init_db()

st.set_page_config(page_title="Shareholder Analytics", page_icon="📈", layout="wide", initial_sidebar_state="expanded")
set_background(BG_IMAGE_PATH)

# --- AUTHENTICATION ---
if 'logged_in' not in st.session_state:
    st.session_state['logged_in'] = False

if not st.session_state['logged_in']:
    auth.login_page()
    st.stop()
    
# --- SIDEBAR MENU ---
logo_path = "logo.png"
if os.path.exists(logo_path):
    if hasattr(st, "logo"):
        st.logo(logo_path)
    else:
        st.sidebar.image(logo_path, use_container_width=True)

menu_options = ["📊 ระบบวิเคราะห์ผู้ถือหุ้น"]
if st.session_state.get('role') == 'admin':
    menu_options.append("⚙️ จัดการผู้ใช้งาน (Admin)")

st.sidebar.title(f"👤 สวัสดีคุณ {st.session_state.get('username', '')}")
page = st.sidebar.radio("เมนูหลัก", menu_options)

if st.sidebar.button("🚪 ออกจากระบบ"):
    st.session_state['logged_in'] = False
    st.session_state['role'] = None
    st.session_state['username'] = None
    st.rerun()

if page == "⚙️ จัดการผู้ใช้งาน (Admin)":
    auth.admin_page()
    st.stop()

st.markdown("""
    <style>
    /* Global CSS Tweaks */
    #MainMenu {visibility: hidden;}
    .stAppDeployButton {display: none;}
    [data-testid="stToolbar"] {visibility: hidden;}
    footer {visibility: hidden;}
    [data-testid="stDataFrame"] {
        background-color: #ffffff;
    }
    .block-container {
        padding-top: 1.5rem;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 20px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        white-space: pre-wrap;
        background-color: transparent;
        border-radius: 4px 4px 0px 0px;
        padding: 10px 16px;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown("""
    <div style='background: linear-gradient(90deg, #1e3a8a 0%, #3b82f6 100%); padding: 25px; border-radius: 12px; margin-bottom: 25px; text-align: center; color: white; box-shadow: 0 4px 6px rgba(0,0,0,0.1);'>
        <h1 style='color: white; margin: 0; font-size: 2.2rem; font-weight: bold;'>📈 ระบบวิเคราะห์และเปรียบเทียบผู้ถือหุ้น</h1>
        <p style='margin: 8px 0 0 0; font-size: 1.1rem; opacity: 0.9;'>Shareholder Analytics Dashboard</p>
    </div>
""", unsafe_allow_html=True)

# Initialize DB on start
db.init_db()

# Tabs
tab1, tab2 = st.tabs(["📈 รายงานเปรียบเทียบ (Report)", "⚙️ อัปโหลดข้อมูล (Upload Data)"])

with tab2:
    st.header("อัพโหลดไฟล์ปิดสมุดทะเบียน (XO) จาก TSD [Excel]")
    st.write("อัปโหลดไฟล์ Excel ที่มีข้อมูลรายชื่อผู้ถือหุ้น ณ วันที่ปิดสมุดทะเบียน (เช่น Sheet data_1)")
    
    uploaded_file = st.file_uploader("เลือกไฟล์ Excel", type=['xlsx'])
    
    # Auto-extract date from filename for default period name
    default_period_name = ""
    if uploaded_file is not None:
        filename = uploaded_file.name
        import re
        # Extract pattern like (17-07-69) or 17-07-69
        match = re.search(r'\(?(\d{2}-\d{2}-\d{2,4})\)?', filename)
        if match:
            # You can also format this to "17 ก.ค. 69" if needed, but keeping the raw date is safer
            default_period_name = match.group(1)
        else:
            # Fallback to filename without extension
            default_period_name = filename.rsplit('.', 1)[0]
            
    period_name = st.text_input("ตั้งชื่อรอบข้อมูลนี้ (เช่น '17-07-69')", value=default_period_name)
    
    selected_sheet = None
    if uploaded_file is not None:
        try:
            xls = pd.ExcelFile(uploaded_file, engine='openpyxl')
            sheet_names = xls.sheet_names
            
            # Try to auto-select a sheet starting with 'data_' if available
            default_index = 0
            for i, name in enumerate(sheet_names):
                if name.startswith('data_'):
                    default_index = i
                    break
                    
            selected_sheet = st.selectbox("เลือกหน้า (Sheet) ที่มีตารางข้อมูลดิบ:", sheet_names, index=default_index)
        except Exception as e:
            st.error(f"ไม่สามารถอ่านไฟล์ได้: {e}")

    if st.button("บันทึกข้อมูล", type="primary"):
        if uploaded_file is not None and period_name and selected_sheet:
            try:
                with st.spinner("กำลังประมวลผลไฟล์..."):
                    df = pd.read_excel(uploaded_file, sheet_name=selected_sheet, engine='openpyxl')
                    
                    # Clean columns for checking
                    cleaned_cols = [str(c).split('\n')[0].strip() for c in df.columns]
                    
                    if 'จำนวนหุ้นทั้งหมด' not in cleaned_cols and 'จำนวนหุ้น' not in cleaned_cols:
                        st.error("ไฟล์ไม่ถูกต้อง: ไม่พบคอลัมน์ 'จำนวนหุ้น' หรือ 'จำนวนหุ้นทั้งหมด' กรุณาตรวจสอบไฟล์อีกครั้ง")
                    else:
                        db.save_shareholders_data(df, period_name)
                        st.success(f"บันทึกข้อมูลรอบ '{period_name}' สำเร็จแล้ว! ({len(df)} รายการ)")
            except Exception as e:
                st.error(f"เกิดข้อผิดพลาด: {e}")
        else:
            st.warning("กรุณากรอกชื่อรอบข้อมูลและเลือกไฟล์ให้ครบถ้วน")
            
    st.markdown("---")
    st.subheader("รอบข้อมูลที่อยู่ในระบบ")
    periods_df = db.get_all_periods()
    if not periods_df.empty:
        st.dataframe(periods_df)
        
        st.markdown("---")
        st.subheader("ลบข้อมูล")
        with st.form("delete_period_form"):
            col1, col2 = st.columns([3, 1])
            with col1:
                period_to_delete = st.selectbox("เลือกรอบข้อมูลที่ต้องการลบ", periods_df['period_name'].tolist())
            with col2:
                st.write("") # Spacing
                st.write("") # Spacing
                submit_delete = st.form_submit_button("🗑️ ลบข้อมูล", type="primary")
                
            if submit_delete:
                if period_to_delete:
                    db.delete_period(period_to_delete)
                    st.success(f"ลบข้อมูลรอบ '{period_to_delete}' เรียบร้อยแล้ว!")
                    time.sleep(1.5)
                    st.rerun()
    else:
        st.info("ยังไม่มีข้อมูลในระบบ")

with tab1:
    st.header("ตารางเปรียบเทียบรายชื่อผู้ถือหุ้น")
    
    periods_df = db.get_all_periods()
    if periods_df.empty:
        st.warning("กรุณาอัปโหลดข้อมูลในแท็บ 'อัปโหลดข้อมูล' ก่อนใช้งาน")
    else:
        all_periods = periods_df['period_name'].tolist()
        
        col1, col2 = st.columns(2)
        with col1:
            selected_periods = st.multiselect("เลือกรอบข้อมูลที่ต้องการเปรียบเทียบ", options=all_periods, default=all_periods[:2] if len(all_periods) >= 2 else all_periods)
        with col2:
            min_shares = st.number_input("จำนวนหุ้นขั้นต่ำที่ต้องการแสดง", min_value=0, value=3000000, step=100000)
            
        if st.button("ประมวลผลเปรียบเทียบ", type="primary"):
            if not selected_periods:
                st.warning("กรุณาเลือกรอบข้อมูลอย่างน้อย 1 รอบ")
            else:
                with st.spinner("กำลังคำนวณ..."):
                    ordered_periods = list(reversed(selected_periods))
                    latest_period = ordered_periods[-1]
                    
                    comp_df = db.get_comparison_data(selected_periods, min_shares, filter_period=latest_period)
                    
                    if comp_df.empty:
                        st.info("ไม่พบข้อมูลที่ตรงกับเงื่อนไขที่กำหนด")
                    else:
                        # Reset index to bring 'full_name' back as a column
                        comp_df = comp_df.reset_index()
                        comp_df = comp_df.rename(columns={'full_name': 'รายชื่อผู้ถือหุ้น'})
                        
                        col_order = ['รายชื่อผู้ถือหุ้น'] + ordered_periods
                        comp_df = comp_df[col_order]
                        
                        # Calculate totals before difference and % to append later
                        totals = {}
                        for p in ordered_periods:
                            col_sum = comp_df[p].sum()
                            total_reg = db.get_total_shares(p)
                            others = total_reg - col_sum
                            totals[p] = {
                                'sum': col_sum,
                                'others': others,
                                'total': total_reg
                            }
                        
                        # Add % column based on latest period
                        total_latest = totals[latest_period]['total']
                        if total_latest > 0:
                            comp_df['%'] = (comp_df[latest_period] / total_latest) * 100
                        else:
                            comp_df['%'] = 0.0
                            
                        # Add Difference column if >= 2 periods
                        diff_col_name = 'เพิ่มขึ้น / (ลดลง)'

                        if len(ordered_periods) >= 2:

                            period_new = ordered_periods[-1] # latest
                            period_old = ordered_periods[-2] # previous
                            
                            comp_df[diff_col_name] = comp_df[period_new].fillna(0) - comp_df[period_old].fillna(0)
                        
                        # Sort by latest period descending
                        comp_df = comp_df.sort_values(by=latest_period, ascending=False)
                        
                        # Add sequence number
                        comp_df.insert(0, 'ลำดับ', range(1, len(comp_df) + 1))
                        
                        # Append total rows
                        row_sum = {'ลำดับ': '', 'รายชื่อผู้ถือหุ้น': 'รวมจำนวนหุ้น'}
                        row_others = {'ลำดับ': '', 'รายชื่อผู้ถือหุ้น': 'อื่นๆ'}
                        row_total = {'ลำดับ': '', 'รายชื่อผู้ถือหุ้น': 'จำนวนหุ้นจดทะเบียน'}
                        
                        for p in ordered_periods:
                            row_sum[p] = totals[p]['sum']
                            row_others[p] = totals[p]['others']
                            row_total[p] = totals[p]['total']
                            
                        row_sum['%'] = (totals[latest_period]['sum'] / total_latest * 100) if total_latest > 0 else 0
                        row_others['%'] = (totals[latest_period]['others'] / total_latest * 100) if total_latest > 0 else 0
                        row_total['%'] = 100.00
                        
                        if len(ordered_periods) >= 2:
                            row_sum[diff_col_name] = None
                            row_others[diff_col_name] = None
                            row_total[diff_col_name] = None
                            
                        summary_df = pd.DataFrame([row_sum, row_others, row_total])
                        comp_df = pd.concat([comp_df, summary_df], ignore_index=True)
                        
                        # Create MultiIndex for columns to match the desired header layout
                        new_cols = []
                        for col in comp_df.columns:
                            if col == 'ลำดับ':
                                new_cols.append(('', 'ลำดับ'))
                            elif col == 'รายชื่อผู้ถือหุ้น':
                                new_cols.append(('', 'รายชื่อผู้ถือหุ้น'))
                            elif col in ordered_periods:
                                new_cols.append((col, 'จำนวนหุ้น'))
                            elif col == '%':
                                new_cols.append((latest_period, '%'))
                            elif col == 'เพิ่มขึ้น / (ลดลง)':
                                top_level = f"{period_old} VS {period_new}" if len(ordered_periods) >= 2 else "เปรียบเทียบ"
                                new_cols.append((top_level, 'จำนวนหุ้น เพิ่มขึ้น / (ลดลง)'))
                            else:
                                new_cols.append(('', col))
                                
                        comp_df.columns = pd.MultiIndex.from_tuples(new_cols)
                        
                        # Formatting dictionary
                        def format_shares(val):
                            # Handles pd.NA, np.nan, string "-", and 0
                            if pd.isna(val) or val == 0 or val == "-":
                                return "-"
                            try:
                                return f"{float(val):,.0f}"
                            except:
                                return str(val)

                        def format_percent(val):
                            if pd.isna(val) or val == 0 or val == "-":
                                return "-"
                            try:
                                return f"{float(val):.4f}%"
                            except:
                                return str(val)
                                
                        def format_diff_val(val):
                            if pd.isna(val) or val == 0 or val == "-":
                                return "-"
                            try:
                                v = float(val)
                                if v > 0:
                                    return f'<div style="display:flex; justify-content:space-between;"><span>▲</span><span>{v:,.0f}</span></div>'
                                elif v < 0:
                                    return f'<div style="display:flex; justify-content:space-between;"><span>▼</span><span>-{abs(v):,.0f}</span></div>'
                                return "-"
                            except:
                                return str(val)

                        format_dict = {}
                        for p in ordered_periods:
                            format_dict[(p, 'จำนวนหุ้น')] = format_shares
                        format_dict[(latest_period, '%')] = format_percent
                        if len(ordered_periods) >= 2:
                            diff_col_tuple = (f"{period_old} VS {period_new}", 'จำนวนหุ้น เพิ่มขึ้น / (ลดลง)')
                            format_dict[diff_col_tuple] = format_diff_val
                            
                        # Set up table styles for HTML rendering (Center headers, Right align numbers, Left align names)
                        table_styles = [
                            {'selector': 'th', 'props': [('text-align', 'center !important'), ('vertical-align', 'middle !important'), ('background-color', '#f0f2f6')]},
                            {'selector': 'td', 'props': [('text-align', 'right')]},
                            {'selector': 'td:nth-child(1), td:nth-child(2)', 'props': [('text-align', 'left')]}
                        ]
                        
                        # Apply style
                        styled_df = comp_df.style.format(formatter=format_dict, na_rep="-").set_table_styles(table_styles)
                        
                        # Hide pandas index
                        if hasattr(styled_df, 'hide'):
                            styled_df = styled_df.hide(axis='index')
                        
                        def bold_summary(row):
                            name = row.iloc[1] 
                            if name in ['รวมจำนวนหุ้น', 'อื่นๆ', 'จำนวนหุ้นจดทะเบียน']:
                                return ['font-weight: bold;'] * len(row)
                            return [''] * len(row)
                        
                        styled_df = styled_df.apply(bold_summary, axis=1)
                        
                        def color_empty(val):
                            if pd.isna(val) or val == 0 or val == "-":
                                return "color: #9ca3af;"
                            return ""
                            
                        cols_to_color = [c for c in comp_df.columns if c[1] in ["จำนวนหุ้น", "%"]]
                        if hasattr(styled_df, "map"):
                            styled_df = styled_df.map(color_empty, subset=cols_to_color)
                        else:
                            styled_df = styled_df.applymap(color_empty, subset=cols_to_color)
                        
                        if len(ordered_periods) >= 2:
                            def color_diff(val):
                                if pd.isna(val) or val == 0 or val == "-":
                                    return 'color: #9ca3af;'
                                try:
                                    v = float(val)
                                    if v > 0:
                                        return 'color: #00cc00;' # Green
                                    elif v < 0:
                                        return 'color: #ff0000;' # Red
                                except:
                                    pass
                                return ''
                            
                            # Use map (Pandas >= 2.1.0) or applymap (older versions)
                            if hasattr(styled_df, 'map'):
                                styled_df = styled_df.map(color_diff, subset=[diff_col_tuple])
                            else:
                                styled_df = styled_df.applymap(color_diff, subset=[diff_col_tuple])
                                
                        
                        import altair as alt
                        
                        dash_tab, table_tab = st.tabs(["📊 แดชบอร์ดสรุป", "📋 ตารางเปรียบเทียบ"])
                        
                        with dash_tab:
                            st.header(f"📌 สรุปข้อมูล ณ {latest_period}")
                            clean_df = comp_df.iloc[:-3].copy() # Exclude summary rows
                            
                            # Metrics Row
                            m1, m2, m3, m4 = st.columns(4)
                            with m1:
                                st.metric("👥 ผู้ถือหุ้นเข้าเกณฑ์", f"{len(clean_df):,.0f} ราย", help=f"ผู้ที่ถือหุ้นตั้งแต่ {min_shares:,.0f} หุ้นขึ้นไป")
                            with m2:
                                st.metric("📈 ผู้ถือหุ้นเพิ่มขึ้น", f"{sum((clean_df[(ordered_periods[-1], 'จำนวนหุ้น')] - clean_df[(ordered_periods[-2], 'จำนวนหุ้น')]) > 0) if len(ordered_periods) >= 2 else 0} ราย")
                            with m3:
                                st.metric("📉 ผู้ถือหุ้นลดลง", f"{sum((clean_df[(ordered_periods[-1], 'จำนวนหุ้น')] - clean_df[(ordered_periods[-2], 'จำนวนหุ้น')]) < 0) if len(ordered_periods) >= 2 else 0} ราย")
                            with m4:
                                st.metric("🏢 หุ้นจดทะเบียนทั้งหมด", f"{totals[latest_period]['total']:,.0f} หุ้น")
                                
                            st.divider()
                            
                            # 1. Pie Chart Top 10
                            top10 = clean_df.head(10)
                            top10_shares = top10[(latest_period, 'จำนวนหุ้น')].sum()
                            total_shares = totals[latest_period]['total']
                            others_shares = total_shares - top10_shares
                            
                            pie_names = top10[('', 'รายชื่อผู้ถือหุ้น')].tolist() + ['อื่นๆ (ผู้ถือหุ้นรายย่อย)']
                            pie_values = top10[(latest_period, 'จำนวนหุ้น')].tolist() + [others_shares]
                            
                            pie_df = pd.DataFrame({'รายชื่อ': pie_names, 'จำนวนหุ้น': pie_values})
                            
                            # Altair Pie Chart
                            base = alt.Chart(pie_df).encode(
                                theta=alt.Theta("จำนวนหุ้น:Q", stack=True),
                                color=alt.Color("รายชื่อ:N", legend=alt.Legend(title="รายชื่อผู้ถือหุ้น", labelLimit=300)),
                                tooltip=['รายชื่อ', alt.Tooltip('จำนวนหุ้น:Q', format=',')]
                            )
                            pie = base.mark_arc(outerRadius=150, innerRadius=70)
                            
                            st.subheader('🍩 สัดส่วนการถือหุ้น 10 รายแรก')
                            st.altair_chart(pie, use_container_width=True)
                            
                            # 2. Increased / Decreased
                            if len(ordered_periods) >= 2:
                                st.divider()
                                col1, col2 = st.columns(2)
                                
                                raw_new = clean_df[(period_new, 'จำนวนหุ้น')]
                                raw_old = clean_df[(period_old, 'จำนวนหุ้น')]
                                raw_diff = raw_new - raw_old
                                
                                inc_mask = raw_diff > 0
                                dec_mask = raw_diff < 0
                                
                                inc_df = clean_df[inc_mask][[ ('', 'รายชื่อผู้ถือหุ้น'), (period_old, 'จำนวนหุ้น'), (period_new, 'จำนวนหุ้น') ]].copy()
                                inc_df['เพิ่มขึ้น'] = raw_diff[inc_mask]
                                inc_df.columns = ['รายชื่อผู้ถือหุ้น', f'เดิม ({period_old})', f'ใหม่ ({period_new})', 'จำนวนที่เพิ่ม']
                                inc_df = inc_df.sort_values(by='จำนวนที่เพิ่ม', ascending=False).reset_index(drop=True)
                                inc_df.index += 1
                                
                                dec_df = clean_df[dec_mask][[ ('', 'รายชื่อผู้ถือหุ้น'), (period_old, 'จำนวนหุ้น'), (period_new, 'จำนวนหุ้น') ]].copy()
                                dec_df['ลดลง'] = raw_diff[dec_mask].abs()
                                dec_df.columns = ['รายชื่อผู้ถือหุ้น', f'เดิม ({period_old})', f'ใหม่ ({period_new})', 'จำนวนที่ลด']
                                dec_df = dec_df.sort_values(by='จำนวนที่ลด', ascending=False).reset_index(drop=True)
                                dec_df.index += 1
                                
                                with col1:
                                    st.subheader(f"🟢 หุ้นเพิ่มขึ้น ({len(inc_df)} ราย)")
                                    if not inc_df.empty:
                                        def style_inc(val):
                                            return 'color: #166534; background-color: #dcfce7; font-weight: bold;'
                                            
                                        inc_styled = inc_df.style.format({
                                            f'เดิม ({period_old})': "{:,.0f}",
                                            f'ใหม่ ({period_new})': "{:,.0f}",
                                            'จำนวนที่เพิ่ม': "{:,.0f}"
                                        })
                                        if hasattr(inc_styled, 'map'):
                                            inc_styled = inc_styled.map(style_inc, subset=['จำนวนที่เพิ่ม'])
                                        else:
                                            inc_styled = inc_styled.applymap(style_inc, subset=['จำนวนที่เพิ่ม'])
                                            
                                        st.dataframe(inc_styled, use_container_width=True)
                                    else:
                                        st.info("ไม่มีรายชื่อที่หุ้นเพิ่มขึ้น")
                                        
                                with col2:
                                    st.subheader(f"🔴 หุ้นลดลง ({len(dec_df)} ราย)")
                                    if not dec_df.empty:
                                        def style_dec(val):
                                            return 'color: #991b1b; background-color: #fee2e2; font-weight: bold;'
                                            
                                        dec_styled = dec_df.style.format({
                                            f'เดิม ({period_old})': "{:,.0f}",
                                            f'ใหม่ ({period_new})': "{:,.0f}",
                                            'จำนวนที่ลด': "{:,.0f}"
                                        })
                                        if hasattr(dec_styled, 'map'):
                                            dec_styled = dec_styled.map(style_dec, subset=['จำนวนที่ลด'])
                                        else:
                                            dec_styled = dec_styled.applymap(style_dec, subset=['จำนวนที่ลด'])
                                            
                                        st.dataframe(dec_styled, use_container_width=True)
                                    else:
                                        st.info("ไม่มีรายชื่อที่หุ้นลดลง")
                        
                        with table_tab:
                            import textwrap
                            # Generate HTML for Web Display
                            html_table = styled_df.to_html()
                            custom_css = textwrap.dedent("""
                            <style>
                            .table-container {
                                max-height: 700px;
                                overflow: auto;
                                border: 1px solid #e2e8f0;
                                border-radius: 12px;
                                margin-bottom: 20px;
                                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
                            }
                            .table-container table {
                                width: 100%;
                                border-collapse: collapse;
                                font-size: 14.5px;
                                background-color: #ffffff;
                            }
                            .table-container th, .table-container td {
                                border: 1px solid #e2e8f0;
                                padding: 10px 14px;
                            }
                            .table-container td {
                                white-space: nowrap;
                            }
                            .table-container td:nth-child(2) {
                                white-space: normal;
                                min-width: 300px;
                            }
                            .table-container th {
                                background-color: #f8fafc !important;
                                color: #1e293b !important;
                                font-weight: 600;
                            }
                            .table-container thead th {
                                position: sticky;
                                z-index: 2;
                            }
                            .table-container thead tr:nth-child(1) th {
                                top: 0;
                            }
                            .table-container thead tr:nth-child(2) th {
                                top: 42px;
                            }
                            .table-container tbody tr:hover {
                                background-color: #f1f5f9;
                            }
                            </style>
                            """)
                            if hasattr(st, 'html'):
                                st.html(custom_css + f'<div class="table-container">\n{html_table}\n</div>')
                            else:
                                html_table = html_table.replace('\n#T_', '\n #T_') # Prevent markdown H1 parsing
                                st.markdown(custom_css + f'<div class="table-container">\n{html_table}\n</div>', unsafe_allow_html=True)
                            
                            # Generate standalone shareable file
                            full_html = textwrap.dedent(f"""
                            <div style="font-family: 'Sarabun', sans-serif;">
                                <div style="display: flex; justify-content: center; align-items: center; margin-bottom: 20px;">
                                    <div style="text-align: center;">
                                        <h2 style="color: #1e3a8a; margin-bottom: 5px;">ตารางเปรียบเทียบรายชื่อผู้ถือหุ้น BEM</h2>
                                        <p style="color: #64748b; margin-top: 0;">ณ วันปิดสมุดทะเบียน {latest_period}</p>
                                    </div>
                                </div>
                                {custom_css}
                                <div class="table-container">
                                    {html_table}
                                </div>
                            </div>
                            """)
                            with open("shareable_report.html", "w", encoding="utf-8") as f:
                                f.write(full_html)
                                
                            st.info("💡 **แชร์ตารางให้ผู้บริหารหรือทีมงาน:** ก๊อปปี้ลิงก์ด้านล่างนี้ส่งให้ได้เลยครับ (คนที่ได้รับลิงก์จะเห็นแค่ตาราง ไม่ต้องล็อกอินครับ)")
                            st.code("https://shareholder-analytics-jirxoabos6p7vvqkq8njfe.streamlit.app/?view=report", language="text")

                            import openpyxl
                            from openpyxl.styles import Font, Alignment, Border, Side
                            from openpyxl.utils import get_column_letter

                            buffer = io.BytesIO()
                            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                # Create a copy and flatten columns to bypass pandas MultiIndex + index=False restriction
                                excel_df = comp_df.copy()
                                excel_df.columns = [str(i) for i in range(len(comp_df.columns))]
                                
                                # Export raw numeric data WITHOUT pandas headers and index
                                excel_df.to_excel(writer, sheet_name='Comparison', index=False, header=False, startrow=4)
                               
                                ws = writer.sheets['Comparison']
                                max_col = len(comp_df.columns)
                                max_row = ws.max_row
                                
                                # 1. Title
                                ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=max_col)
                                title_cell = ws.cell(row=1, column=1)
                                title_cell.value = f"ตารางเปรียบเทียบรายชื่อผู้ถือหุ้น\nณ วันปิดสมุดทะเบียน"
                                title_cell.font = Font(name='Sarabun', size=16, bold=True)
                                title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                
                                # 2. Write headers manually
                                current_top = None
                                merge_start = None
                                
                                for i, (top_lvl, bot_lvl) in enumerate(comp_df.columns, start=1):
                                    ws.cell(row=4, column=i, value=bot_lvl)
                                    ws.cell(row=3, column=i, value=top_lvl)
                                    
                                    # Horizontal merge tracking for top level
                                    if top_lvl != current_top:
                                        if current_top != "" and current_top is not None and merge_start < i - 1:
                                            ws.merge_cells(start_row=3, start_column=merge_start, end_row=3, end_column=i - 1)
                                        current_top = top_lvl
                                        merge_start = i
                                        
                                # Handle last horizontal merge
                                if current_top != "" and current_top is not None and merge_start < max_col:
                                    ws.merge_cells(start_row=3, start_column=merge_start, end_row=3, end_column=max_col)
                                    
                                # Vertical merge for empty top levels
                                for i, (top_lvl, bot_lvl) in enumerate(comp_df.columns, start=1):
                                    if top_lvl == "":
                                        ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)
                                        ws.cell(row=3, column=i, value=bot_lvl)
                                
                                # Setup borders
                                thin = Side(border_style="thin", color="000000")
                                border = Border(top=thin, left=thin, right=thin, bottom=thin)
                                
                                # Apply styles to all cells (borders and font)
                                for r_idx in range(3, max_row + 1):
                                    for c_idx in range(1, max_col + 1):
                                        cell = ws.cell(row=r_idx, column=c_idx)
                                        cell.border = border
                                        cell.font = Font(name='Sarabun', size=11)
                                        cell.alignment = Alignment(vertical='center')
                                
                                # Header styles
                                for r_idx in range(3, 5):
                                    for c_idx in range(1, max_col + 1):
                                        cell = ws.cell(row=r_idx, column=c_idx)
                                        cell.font = Font(name='Sarabun', size=12, bold=True)
                                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                        
                                # Data styles (row 5 onwards)
                                for r_idx in range(5, max_row + 1):
                                    ws.cell(row=r_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
                                    ws.cell(row=r_idx, column=2).alignment = Alignment(horizontal='left', vertical='center')
                                    
                                    for c_idx in range(3, max_col + 1):
                                        cell = ws.cell(row=r_idx, column=c_idx)
                                        cell.alignment = Alignment(horizontal='right', vertical='center')
                                        col_name = comp_df.columns[c_idx-1][1]
                                        
                                        if col_name == '%':
                                            cell.number_format = '0.000%'
                                            if isinstance(cell.value, (int, float)):
                                                cell.value = cell.value / 100
                                        elif col_name == 'จำนวนหุ้น':
                                            cell.number_format = '#,##0'
                                        elif 'เพิ่มขึ้น' in col_name or 'ลดลง' in col_name:
                                            # Excel conditional formatting via custom number format
                                            cell.number_format = '[Color10]▲* #,##0;[Red]▼* -#,##0;"-"'
                                
                                # Summary rows bolding (last 3 rows)
                                for r_idx in range(max_row - 2, max_row + 1):
                                    for c_idx in range(1, max_col + 1):
                                        ws.cell(row=r_idx, column=c_idx).font = Font(name='Sarabun', size=11, bold=True)
                                        
                                # Column widths
                                ws.column_dimensions['A'].width = 8
                                ws.column_dimensions['B'].width = 50
                                for c_idx in range(3, max_col + 1):
                                    col_letter = get_column_letter(c_idx)
                                    ws.column_dimensions[col_letter].width = 20
                            
                            st.download_button(
                                label="📥 ดาวน์โหลดเป็น Excel",
                                data=buffer.getvalue(),
                                file_name="shareholder_comparison.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
